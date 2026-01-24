"""
Report Generator Service
Generate PDF, Excel, and CSV reports
"""

import io
import csv
import logging
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel
from enum import Enum

logger = logging.getLogger(__name__)


class ReportType(str, Enum):
    """Available report types"""
    REVENUE_SUMMARY = "revenue_summary"
    TECHNICIAN_PERFORMANCE = "technician_performance"
    JOB_ANALYSIS = "job_analysis"
    CUSTOMER_REPORT = "customer_report"
    AR_AGING = "ar_aging"
    WEEKLY_SUMMARY = "weekly_summary"
    MONTHLY_PL = "monthly_pl"


class ReportFormat(str, Enum):
    """Report output formats"""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"


class ReportSchedule(str, Enum):
    """Report schedule options"""
    DAILY = "daily"
    WEEKLY_MONDAY = "monday_8am"
    WEEKLY_FRIDAY = "friday_5pm"
    MONTHLY_FIRST = "monthly_first"
    MONTHLY_LAST = "monthly_last"


class ScheduledReport(BaseModel):
    """Scheduled report configuration"""
    schedule_id: str
    business_id: str
    report_type: ReportType
    schedule: ReportSchedule
    recipients: List[str]
    format: ReportFormat = ReportFormat.PDF
    is_active: bool = True
    last_sent: Optional[datetime] = None
    created_at: datetime = datetime.utcnow()


class ReportGenerator:
    """Service for generating business reports"""

    def __init__(self, db: AsyncIOMotorDatabase, business_id: str):
        self.db = db
        self.business_id = business_id

    async def generate_report(
        self,
        report_type: ReportType,
        format: ReportFormat,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """Generate a report of the specified type and format"""
        if not start_date:
            start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
        if not end_date:
            end_date = datetime.utcnow()

        # Get report data
        if report_type == ReportType.REVENUE_SUMMARY:
            data = await self._get_revenue_summary_data(start_date, end_date)
        elif report_type == ReportType.TECHNICIAN_PERFORMANCE:
            data = await self._get_tech_performance_data(start_date, end_date)
        elif report_type == ReportType.JOB_ANALYSIS:
            data = await self._get_job_analysis_data(start_date, end_date)
        elif report_type == ReportType.CUSTOMER_REPORT:
            data = await self._get_customer_report_data(start_date, end_date)
        elif report_type == ReportType.AR_AGING:
            data = await self._get_ar_aging_data()
        elif report_type == ReportType.WEEKLY_SUMMARY:
            data = await self._get_weekly_summary_data()
        elif report_type == ReportType.MONTHLY_PL:
            data = await self._get_monthly_pl_data(start_date, end_date)
        else:
            raise ValueError(f"Unknown report type: {report_type}")

        # Generate output
        if format == ReportFormat.CSV:
            content = self._generate_csv(data, report_type)
            content_type = "text/csv"
            extension = "csv"
        elif format == ReportFormat.EXCEL:
            content = self._generate_excel(data, report_type)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"
        elif format == ReportFormat.PDF:
            content = await self._generate_pdf(data, report_type)
            content_type = "application/pdf"
            extension = "pdf"
        else:
            raise ValueError(f"Unknown format: {format}")

        # Generate filename
        date_str = datetime.utcnow().strftime("%Y%m%d")
        filename = f"{report_type.value}_{date_str}.{extension}"

        return {
            "content": content,
            "content_type": content_type,
            "filename": filename,
            "report_type": report_type,
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            }
        }

    # Data Gathering Methods

    async def _get_revenue_summary_data(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get revenue summary data"""
        # Get completed jobs
        pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "status": "completed",
                    "completed_at": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$group": {
                    "_id": {
                        "year": {"$year": "$completed_at"},
                        "month": {"$month": "$completed_at"},
                        "day": {"$dayOfMonth": "$completed_at"}
                    },
                    "revenue": {"$sum": "$total"},
                    "job_count": {"$sum": 1},
                    "avg_job_value": {"$avg": "$total"}
                }
            },
            {"$sort": {"_id": 1}}
        ]

        daily_data = await self.db.hvac_quotes.aggregate(pipeline).to_list(length=366)

        # Job type breakdown
        type_pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "status": "completed",
                    "completed_at": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$group": {
                    "_id": "$job_type",
                    "revenue": {"$sum": "$total"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"revenue": -1}}
        ]

        by_type = await self.db.hvac_quotes.aggregate(type_pipeline).to_list(length=50)

        # Get business info
        business = await self.db.businesses.find_one({"business_id": self.business_id})

        total_revenue = sum(d["revenue"] for d in daily_data)
        total_jobs = sum(d["job_count"] for d in daily_data)

        return {
            "title": "Revenue Summary Report",
            "business_name": business.get("name", ""),
            "period": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "total_revenue": round(total_revenue, 2),
                "total_jobs": total_jobs,
                "avg_job_value": round(total_revenue / total_jobs, 2) if total_jobs > 0 else 0
            },
            "daily_data": [
                {
                    "date": f"{d['_id']['year']}-{d['_id']['month']:02d}-{d['_id']['day']:02d}",
                    "revenue": round(d["revenue"], 2),
                    "jobs": d["job_count"],
                    "avg_value": round(d["avg_job_value"], 2)
                }
                for d in daily_data
            ],
            "by_type": [
                {
                    "type": t["_id"] or "Other",
                    "revenue": round(t["revenue"], 2),
                    "count": t["count"],
                    "percentage": round(t["revenue"] / total_revenue * 100, 1) if total_revenue > 0 else 0
                }
                for t in by_type
            ]
        }

    async def _get_tech_performance_data(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get technician performance data"""
        pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "status": "completed",
                    "completed_at": {"$gte": start_date, "$lte": end_date},
                    "assigned_tech_id": {"$exists": True, "$ne": None}
                }
            },
            {
                "$group": {
                    "_id": "$assigned_tech_id",
                    "jobs_completed": {"$sum": 1},
                    "revenue": {"$sum": "$total"},
                    "total_hours": {"$sum": "$actual_hours"}
                }
            },
            {"$sort": {"revenue": -1}}
        ]

        tech_data = await self.db.hvac_quotes.aggregate(pipeline).to_list(length=100)

        # Get tech names
        tech_ids = [t["_id"] for t in tech_data]
        techs = await self.db.staff.find({"staff_id": {"$in": tech_ids}}).to_list(length=100)
        tech_map = {t["staff_id"]: t for t in techs}

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Technician Performance Report",
            "business_name": business.get("name", ""),
            "period": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            "generated_at": datetime.utcnow().isoformat(),
            "technicians": [
                {
                    "name": f"{tech_map.get(t['_id'], {}).get('first_name', '')} {tech_map.get(t['_id'], {}).get('last_name', '')}".strip() or "Unknown",
                    "jobs_completed": t["jobs_completed"],
                    "revenue": round(t["revenue"], 2),
                    "hours_worked": round(t.get("total_hours", 0), 1),
                    "revenue_per_hour": round(t["revenue"] / t["total_hours"], 2) if t.get("total_hours", 0) > 0 else 0
                }
                for t in tech_data
            ]
        }

    async def _get_job_analysis_data(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get job analysis data"""
        # Status breakdown
        status_pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "created_at": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$group": {
                    "_id": "$status",
                    "count": {"$sum": 1},
                    "total_value": {"$sum": "$total"}
                }
            }
        ]

        by_status = await self.db.hvac_quotes.aggregate(status_pipeline).to_list(length=20)

        # Conversion funnel
        total_created = sum(s["count"] for s in by_status)
        completed = next((s for s in by_status if s["_id"] == "completed"), {"count": 0})
        cancelled = next((s for s in by_status if s["_id"] == "cancelled"), {"count": 0})

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Job Analysis Report",
            "business_name": business.get("name", ""),
            "period": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            "generated_at": datetime.utcnow().isoformat(),
            "status_breakdown": [
                {
                    "status": s["_id"],
                    "count": s["count"],
                    "value": round(s["total_value"], 2)
                }
                for s in by_status
            ],
            "conversion": {
                "total_jobs": total_created,
                "completed": completed["count"],
                "cancelled": cancelled["count"],
                "completion_rate": round(completed["count"] / total_created * 100, 1) if total_created > 0 else 0
            }
        }

    async def _get_customer_report_data(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get customer report data"""
        # New customers
        new_customers = await self.db.clients.count_documents({
            "business_id": self.business_id,
            "created_at": {"$gte": start_date, "$lte": end_date}
        })

        # Top customers by revenue
        top_pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "status": "completed",
                    "completed_at": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$group": {
                    "_id": "$client_id",
                    "revenue": {"$sum": "$total"},
                    "job_count": {"$sum": 1}
                }
            },
            {"$sort": {"revenue": -1}},
            {"$limit": 20}
        ]

        top_customers = await self.db.hvac_quotes.aggregate(top_pipeline).to_list(length=20)

        # Get customer details
        customer_ids = [c["_id"] for c in top_customers]
        customers = await self.db.clients.find({"client_id": {"$in": customer_ids}}).to_list(length=20)
        customer_map = {c["client_id"]: c for c in customers}

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Customer Report",
            "business_name": business.get("name", ""),
            "period": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            "generated_at": datetime.utcnow().isoformat(),
            "new_customers": new_customers,
            "top_customers": [
                {
                    "name": f"{customer_map.get(c['_id'], {}).get('first_name', '')} {customer_map.get(c['_id'], {}).get('last_name', '')}".strip() or "Unknown",
                    "email": customer_map.get(c["_id"], {}).get("email", ""),
                    "revenue": round(c["revenue"], 2),
                    "jobs": c["job_count"]
                }
                for c in top_customers
            ]
        }

    async def _get_ar_aging_data(self) -> Dict[str, Any]:
        """Get accounts receivable aging data"""
        now = datetime.utcnow()

        invoices = await self.db.invoices.find({
            "business_id": self.business_id,
            "status": {"$in": ["sent", "overdue"]}
        }).to_list(length=1000)

        buckets = {
            "current": [],
            "30_days": [],
            "60_days": [],
            "90_days": [],
            "120_plus": []
        }

        for inv in invoices:
            invoice_date = inv.get("invoice_date")
            if not invoice_date:
                continue

            if isinstance(invoice_date, str):
                invoice_date = datetime.fromisoformat(invoice_date.replace("Z", "+00:00"))

            days_old = (now - invoice_date).days
            amount = inv.get("amount_due", 0)

            entry = {
                "invoice_number": inv.get("invoice_number", ""),
                "customer": inv.get("client_name", ""),
                "amount": round(amount, 2),
                "date": invoice_date.strftime("%Y-%m-%d"),
                "days_outstanding": days_old
            }

            if days_old <= 30:
                buckets["current"].append(entry)
            elif days_old <= 60:
                buckets["30_days"].append(entry)
            elif days_old <= 90:
                buckets["60_days"].append(entry)
            elif days_old <= 120:
                buckets["90_days"].append(entry)
            else:
                buckets["120_plus"].append(entry)

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Accounts Receivable Aging Report",
            "business_name": business.get("name", ""),
            "generated_at": datetime.utcnow().isoformat(),
            "buckets": {
                k: {
                    "invoices": v,
                    "total": sum(i["amount"] for i in v),
                    "count": len(v)
                }
                for k, v in buckets.items()
            },
            "grand_total": sum(sum(i["amount"] for i in v) for v in buckets.values())
        }

    async def _get_weekly_summary_data(self) -> Dict[str, Any]:
        """Get weekly summary data"""
        now = datetime.utcnow()
        week_start = now - timedelta(days=7)

        # Jobs this week
        jobs = await self.db.hvac_quotes.find({
            "business_id": self.business_id,
            "completed_at": {"$gte": week_start, "$lte": now}
        }).to_list(length=500)

        total_revenue = sum(j.get("total", 0) for j in jobs)
        completed = len([j for j in jobs if j.get("status") == "completed"])

        # New customers
        new_customers = await self.db.clients.count_documents({
            "business_id": self.business_id,
            "created_at": {"$gte": week_start}
        })

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Weekly Summary Report",
            "business_name": business.get("name", ""),
            "week_of": week_start.strftime("%B %d, %Y"),
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "total_revenue": round(total_revenue, 2),
                "jobs_completed": completed,
                "new_customers": new_customers,
                "avg_job_value": round(total_revenue / completed, 2) if completed > 0 else 0
            }
        }

    async def _get_monthly_pl_data(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get monthly P&L data"""
        # Revenue
        revenue_pipeline = [
            {
                "$match": {
                    "business_id": self.business_id,
                    "status": "completed",
                    "completed_at": {"$gte": start_date, "$lte": end_date}
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total"},
                    "total_cost": {"$sum": "$cost"}
                }
            }
        ]

        revenue_data = await self.db.hvac_quotes.aggregate(revenue_pipeline).to_list(length=1)

        total_revenue = revenue_data[0]["total_revenue"] if revenue_data else 0
        total_cost = revenue_data[0].get("total_cost", 0) if revenue_data else 0
        gross_profit = total_revenue - total_cost

        business = await self.db.businesses.find_one({"business_id": self.business_id})

        return {
            "title": "Monthly Profit & Loss Report",
            "business_name": business.get("name", ""),
            "period": f"{start_date.strftime('%B %Y')}",
            "generated_at": datetime.utcnow().isoformat(),
            "income": {
                "service_revenue": round(total_revenue, 2)
            },
            "expenses": {
                "cost_of_goods": round(total_cost, 2)
            },
            "gross_profit": round(gross_profit, 2),
            "gross_margin": round(gross_profit / total_revenue * 100, 1) if total_revenue > 0 else 0
        }

    # Output Generation Methods

    def _generate_csv(self, data: Dict[str, Any], report_type: ReportType) -> bytes:
        """Generate CSV output"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header info
        writer.writerow([data.get("title", "Report")])
        writer.writerow([f"Generated: {data.get('generated_at', '')}"])
        writer.writerow([])

        if report_type == ReportType.REVENUE_SUMMARY:
            # Summary
            writer.writerow(["Summary"])
            writer.writerow(["Total Revenue", data["summary"]["total_revenue"]])
            writer.writerow(["Total Jobs", data["summary"]["total_jobs"]])
            writer.writerow(["Avg Job Value", data["summary"]["avg_job_value"]])
            writer.writerow([])

            # Daily data
            writer.writerow(["Daily Breakdown"])
            writer.writerow(["Date", "Revenue", "Jobs", "Avg Value"])
            for row in data.get("daily_data", []):
                writer.writerow([row["date"], row["revenue"], row["jobs"], row["avg_value"]])

        elif report_type == ReportType.TECHNICIAN_PERFORMANCE:
            writer.writerow(["Name", "Jobs Completed", "Revenue", "Hours Worked", "Revenue/Hour"])
            for tech in data.get("technicians", []):
                writer.writerow([
                    tech["name"],
                    tech["jobs_completed"],
                    tech["revenue"],
                    tech["hours_worked"],
                    tech["revenue_per_hour"]
                ])

        elif report_type == ReportType.AR_AGING:
            for bucket_name, bucket_data in data.get("buckets", {}).items():
                writer.writerow([f"{bucket_name.upper()} (Total: ${bucket_data['total']})"])
                writer.writerow(["Invoice #", "Customer", "Amount", "Date", "Days Outstanding"])
                for inv in bucket_data["invoices"]:
                    writer.writerow([
                        inv["invoice_number"],
                        inv["customer"],
                        inv["amount"],
                        inv["date"],
                        inv["days_outstanding"]
                    ])
                writer.writerow([])

        else:
            # Generic handling
            writer.writerow(["Report Data"])
            for key, value in data.items():
                if isinstance(value, (str, int, float)):
                    writer.writerow([key, value])

        return output.getvalue().encode('utf-8')

    def _generate_excel(self, data: Dict[str, Any], report_type: ReportType) -> bytes:
        """Generate Excel output using openpyxl if available, otherwise CSV"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Header styling
            header_font = Font(bold=True, size=14)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

            # Title
            ws['A1'] = data.get("title", "Report")
            ws['A1'].font = Font(bold=True, size=16)

            ws['A2'] = f"Generated: {data.get('generated_at', '')}"
            ws['A3'] = f"Period: {data.get('period', '')}"

            row = 5

            if report_type == ReportType.REVENUE_SUMMARY:
                # Summary section
                ws.cell(row=row, column=1, value="Summary").font = header_font
                row += 1
                ws.cell(row=row, column=1, value="Total Revenue")
                ws.cell(row=row, column=2, value=data["summary"]["total_revenue"])
                row += 1
                ws.cell(row=row, column=1, value="Total Jobs")
                ws.cell(row=row, column=2, value=data["summary"]["total_jobs"])
                row += 1
                ws.cell(row=row, column=1, value="Avg Job Value")
                ws.cell(row=row, column=2, value=data["summary"]["avg_job_value"])
                row += 2

                # Daily data
                ws.cell(row=row, column=1, value="Daily Breakdown").font = header_font
                row += 1
                headers = ["Date", "Revenue", "Jobs", "Avg Value"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill

                row += 1
                for daily in data.get("daily_data", []):
                    ws.cell(row=row, column=1, value=daily["date"])
                    ws.cell(row=row, column=2, value=daily["revenue"])
                    ws.cell(row=row, column=3, value=daily["jobs"])
                    ws.cell(row=row, column=4, value=daily["avg_value"])
                    row += 1

            elif report_type == ReportType.TECHNICIAN_PERFORMANCE:
                headers = ["Name", "Jobs Completed", "Revenue", "Hours Worked", "Revenue/Hour"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill

                row += 1
                for tech in data.get("technicians", []):
                    ws.cell(row=row, column=1, value=tech["name"])
                    ws.cell(row=row, column=2, value=tech["jobs_completed"])
                    ws.cell(row=row, column=3, value=tech["revenue"])
                    ws.cell(row=row, column=4, value=tech["hours_worked"])
                    ws.cell(row=row, column=5, value=tech["revenue_per_hour"])
                    row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            # Fall back to CSV if openpyxl not available
            logger.warning("openpyxl not available, generating CSV instead")
            return self._generate_csv(data, report_type)

    async def _generate_pdf(self, data: Dict[str, Any], report_type: ReportType) -> bytes:
        """Generate PDF output using WeasyPrint if available"""
        try:
            from weasyprint import HTML

            # Build HTML content
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: #4CAF50; }}
                    h2 {{ color: #333; border-bottom: 2px solid #4CAF50; padding-bottom: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                    th {{ background-color: #4CAF50; color: white; padding: 10px; text-align: left; }}
                    td {{ border: 1px solid #ddd; padding: 8px; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .summary {{ background-color: #f5f5f5; padding: 20px; border-radius: 8px; margin: 20px 0; }}
                    .summary-item {{ display: inline-block; margin-right: 40px; }}
                    .summary-value {{ font-size: 24px; font-weight: bold; color: #4CAF50; }}
                    .summary-label {{ color: #666; }}
                    .meta {{ color: #666; font-size: 12px; }}
                </style>
            </head>
            <body>
                <h1>{data.get('title', 'Report')}</h1>
                <p class="meta">
                    {data.get('business_name', '')}<br>
                    Period: {data.get('period', '')}<br>
                    Generated: {data.get('generated_at', '')}
                </p>
            """

            if report_type == ReportType.REVENUE_SUMMARY:
                summary = data.get("summary", {})
                html_content += f"""
                <div class="summary">
                    <div class="summary-item">
                        <div class="summary-value">${summary.get('total_revenue', 0):,.2f}</div>
                        <div class="summary-label">Total Revenue</div>
                    </div>
                    <div class="summary-item">
                        <div class="summary-value">{summary.get('total_jobs', 0)}</div>
                        <div class="summary-label">Total Jobs</div>
                    </div>
                    <div class="summary-item">
                        <div class="summary-value">${summary.get('avg_job_value', 0):,.2f}</div>
                        <div class="summary-label">Avg Job Value</div>
                    </div>
                </div>

                <h2>Revenue by Job Type</h2>
                <table>
                    <tr><th>Type</th><th>Revenue</th><th>Jobs</th><th>%</th></tr>
                """
                for t in data.get("by_type", []):
                    html_content += f"""
                    <tr>
                        <td>{t['type']}</td>
                        <td>${t['revenue']:,.2f}</td>
                        <td>{t['count']}</td>
                        <td>{t['percentage']}%</td>
                    </tr>
                    """
                html_content += "</table>"

            elif report_type == ReportType.TECHNICIAN_PERFORMANCE:
                html_content += """
                <h2>Technician Performance</h2>
                <table>
                    <tr><th>Name</th><th>Jobs</th><th>Revenue</th><th>Hours</th><th>$/Hour</th></tr>
                """
                for tech in data.get("technicians", []):
                    html_content += f"""
                    <tr>
                        <td>{tech['name']}</td>
                        <td>{tech['jobs_completed']}</td>
                        <td>${tech['revenue']:,.2f}</td>
                        <td>{tech['hours_worked']}</td>
                        <td>${tech['revenue_per_hour']:,.2f}</td>
                    </tr>
                    """
                html_content += "</table>"

            elif report_type == ReportType.AR_AGING:
                html_content += f"""
                <div class="summary">
                    <div class="summary-item">
                        <div class="summary-value">${data.get('grand_total', 0):,.2f}</div>
                        <div class="summary-label">Total Outstanding</div>
                    </div>
                </div>
                """

                for bucket_name, bucket_data in data.get("buckets", {}).items():
                    if bucket_data["count"] > 0:
                        html_content += f"""
                        <h2>{bucket_name.replace('_', ' ').title()} (${bucket_data['total']:,.2f})</h2>
                        <table>
                            <tr><th>Invoice #</th><th>Customer</th><th>Amount</th><th>Date</th><th>Days</th></tr>
                        """
                        for inv in bucket_data["invoices"][:10]:
                            html_content += f"""
                            <tr>
                                <td>{inv['invoice_number']}</td>
                                <td>{inv['customer']}</td>
                                <td>${inv['amount']:,.2f}</td>
                                <td>{inv['date']}</td>
                                <td>{inv['days_outstanding']}</td>
                            </tr>
                            """
                        html_content += "</table>"

            html_content += "</body></html>"

            pdf = HTML(string=html_content).write_pdf()
            return pdf

        except ImportError:
            logger.warning("WeasyPrint not available, generating HTML instead")
            return f"<html><body><h1>{data.get('title')}</h1><pre>{data}</pre></body></html>".encode('utf-8')

    # Scheduled Reports

    async def get_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Get all scheduled reports for the business"""
        schedules = await self.db.report_schedules.find({
            "business_id": self.business_id
        }).to_list(length=100)

        return schedules

    async def create_schedule(
        self,
        report_type: ReportType,
        schedule: ReportSchedule,
        recipients: List[str],
        format: ReportFormat = ReportFormat.PDF
    ) -> Dict[str, Any]:
        """Create a scheduled report"""
        import uuid

        schedule_doc = {
            "schedule_id": str(uuid.uuid4()),
            "business_id": self.business_id,
            "report_type": report_type.value,
            "schedule": schedule.value,
            "recipients": recipients,
            "format": format.value,
            "is_active": True,
            "last_sent": None,
            "created_at": datetime.utcnow()
        }

        await self.db.report_schedules.insert_one(schedule_doc)
        return schedule_doc

    async def delete_schedule(self, schedule_id: str) -> bool:
        """Delete a scheduled report"""
        result = await self.db.report_schedules.delete_one({
            "schedule_id": schedule_id,
            "business_id": self.business_id
        })
        return result.deleted_count > 0
